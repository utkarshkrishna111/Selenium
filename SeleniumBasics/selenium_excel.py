import time
from selenium import webdriver
#Chrome Driver
from selenium.webdriver.chrome.service import Service
#By contains locators
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl

#Runs the chrome browser in headless mode, ie, without opening te browser on screen
chrome_options = webdriver.ChromeOptions()
#chrome_options.add_argument("headless")
chrome_options.add_argument("--start-maximized")

#To Ignore the warning while opening a website
chrome_options.add_argument("--ignore-certificate-errors")

#Chrome driver Service Selenium ->160 -> chrome driver
service_obj = Service(r"C:\Users\utkar\Downloads\chromedriver.exe")
driver = webdriver.Chrome(service=service_obj, options = chrome_options)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

driver.implicitly_wait(10)

driver.find_element(By.ID,"downloadButton").click()

#edit the excel
file_path = r"C:\Users\utkar\Downloads\download.xlsx"
file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
fruit_name = 'Apple'
file_input.send_keys(file_path)

toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait = WebDriverWait(driver,5)
wait.until(expected_conditions.visibility_of_element_located(toast_locator))

print(driver.find_element(*toast_locator).text)

pricecolumn = (driver.find_element
               (By.XPATH,"//div[text()='Price']").
                get_attribute("data-column-id"))

actual_price = driver.find_element(By.XPATH,
                    "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+pricecolumn+"-undefined']").text
print(actual_price)


book = openpyxl.load_workbook(file_path)

sheet = book.active

cell = sheet.cell(row=1,column=2)
print(cell.value)

print((sheet.cell(row=2,column=2).value))

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A5'].value)

dict = {}
price_column = 0
apple_column = 0
apple_row = 0
for i in range(1,sheet.max_column+1):
    if sheet.cell(row=1,column=i).value == 'price':
        price_column = i
    for j in range(1, sheet.max_row + 1):
        if sheet.cell(row=j, column=i).value == 'Apple':
            apple_column = i
            apple_row = j

print((sheet.cell(row=apple_row, column=apple_column).value))

print((sheet.cell(row=apple_row, column=price_column).value))

apple_price = (sheet.cell(row=apple_row, column=price_column).value)

assert apple_price == int(actual_price)

print((sheet.cell(row=1,column=price_column).value))
















time.sleep(5)